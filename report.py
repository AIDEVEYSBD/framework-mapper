#!/usr/bin/env python3
"""
Framework Mapping Report Generator
Consumes the output from the LLM Framework Overlap Analyzer and framework definition files
to produce a formatted, human-readable Excel framework mapping report with executive summary.
"""
import sys
import logging
import pandas as pd
from itertools import groupby
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# === Helper functions ===

def load_excel_file(file_path, sheet_name=0):
    logging.info(f"Loading data from {file_path}")
    try:
        file_path = Path(file_path)
        if file_path.suffix.lower() in ['.xlsx', '.xls']:
            # Using openpyxl for .xlsx is generally preferred
            return pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl')
        else:
            return pd.read_csv(file_path)
    except Exception as e:
        logging.error(f"Error loading file {file_path}: {e}")
        raise

def map_framework_columns(framework_df):
    logging.info("Mapping source framework columns...")
    required_cols = ['ID', 'Domain', 'Sub-Domain', 'Controls']
    if not all(col in framework_df.columns for col in required_cols):
        missing = set(required_cols) - set(framework_df.columns)
        msg = f"Framework file missing required columns: {missing}."
        logging.error(msg)
        return None, msg

    # Rename to standard internal names for Framework A
    mapped = framework_df.rename(columns={
        'ID': 'ID',
        'Domain': 'Framework A Control Domain',
        'Sub-Domain': 'Framework A Control Sub-Domain',
        'Controls': 'Framework A Control Statement'
    })
    mapped['Framework A Control ID'] = mapped['ID']
    final_cols = ['ID', 'Framework A Control ID', 'Framework A Control Domain',
                  'Framework A Control Sub-Domain', 'Framework A Control Statement']
    logging.info("Framework columns mapped successfully.")
    return mapped[final_cols], None

def normalize_column_names(df):
    """
    Normalize column names to handle both old validator format and new analyzer format.
    Maps old column names to new expected names for backward compatibility.
    """
    logging.info(f"Input columns: {list(df.columns)}")
    
    column_mapping = {
        # Old validator format -> New analyzer format
        'match_type': 'equivalence_type',
        'validation_justification': 'mapping_justification', 
        'gaps_identified': 'distinct_concepts',
        'audit_notes': 'mapping_notes'
    }
    
    # Check if we're dealing with old format
    old_format_detected = any(old_col in df.columns for old_col in column_mapping.keys())
    if old_format_detected:
        logging.info("Old validator format detected - applying column normalization")
    else:
        logging.info("New analyzer format detected")
    
    # Create a copy and rename columns
    normalized_df = df.copy()
    normalized_df = normalized_df.rename(columns=column_mapping)
    
    # Add missing columns with default values if they don't exist
    if 'overlapping_concepts' not in normalized_df.columns:
        # Try to extract overlapping concepts from other fields or create empty
        normalized_df['overlapping_concepts'] = ''
        logging.info("Added empty 'overlapping_concepts' column")
    
    # Convert old match types to new equivalence types
    if 'equivalence_type' in normalized_df.columns and old_format_detected:
        type_mapping = {
            'full': 'equivalent',
            'partial': 'overlapping', 
            'no_match': 'distinct'
        }
        old_values = normalized_df['equivalence_type'].value_counts()
        logging.info(f"Converting equivalence types - found: {dict(old_values)}")
        normalized_df['equivalence_type'] = normalized_df['equivalence_type'].replace(type_mapping)
        new_values = normalized_df['equivalence_type'].value_counts()
        logging.info(f"After conversion: {dict(new_values)}")
    
    logging.info(f"Normalized columns: {list(normalized_df.columns)}")
    return normalized_df

def process_mapped_data(mapped_df):
    """
    Transforms the framework overlap analyzer's output into individual rows for the mapping report.
    - If a Framework A control has any 'equivalent' or 'overlapping' mappings, it shows only those.
    - If a Framework A control has only 'distinct' mappings, it shows one 'No Mapping' row.
    """
    logging.info("Processing framework overlap data into mapping report format.")
    
    # First normalize column names for backward compatibility
    mapped_df = normalize_column_names(mapped_df)
    
    required_cols = ['source_id', 'source_text', 'target_id', 'target_text',
                     'equivalence_type', 'mapping_justification', 'overlapping_concepts',
                     'distinct_concepts', 'mapping_notes']

    if not all(col in mapped_df.columns for col in required_cols):
        missing = set(required_cols) - set(mapped_df.columns)
        msg = f"Missing required columns in overlap analysis file after normalization: {missing}. Available: {list(mapped_df.columns)}"
        logging.error(msg)
        return None, msg

    # Status map for framework overlap types (handles both old and new formats)
    status_map = {
        # New format
        'equivalent': 'Direct Mapping',
        'overlapping': 'Partial Mapping',
        'distinct': 'No Mapping',
        # Old format (for backward compatibility)
        'full': 'Direct Mapping',
        'partial': 'Partial Mapping', 
        'no_match': 'No Mapping'
    }
    
    processed_data = []
    
    # Group by source_id to process each Framework A control individually
    for source_id, group in mapped_df.groupby('source_id'):
        # Sort by mapping quality (equivalent > overlapping > distinct)
        mapping_order = {'equivalent': 0, 'overlapping': 1, 'distinct': 2}
        group = group.copy()
        group['mapping_order'] = group['equivalence_type'].map(mapping_order)
        group = group.sort_values('mapping_order')
        
        # Check if there are any good mappings (equivalent or overlapping)
        has_good_mappings = any(group['equivalence_type'].isin(['equivalent', 'overlapping']))
        
        if has_good_mappings:
            # Filter to show only the top 3 best mappings
            good_mappings = group[group['equivalence_type'].isin(['equivalent', 'overlapping'])].head(3)
            
            for _, row in good_mappings.iterrows():
                # Build detailed mapping explanation from analysis results
                explanation_parts = []
                if pd.notna(row['mapping_justification']) and str(row['mapping_justification']).strip():
                    explanation_parts.append(str(row['mapping_justification']))
                
                # Handle overlapping_concepts (may be empty from normalization)
                if pd.notna(row['overlapping_concepts']) and str(row['overlapping_concepts']).strip():
                    explanation_parts.append(f"\n\nShared Concepts: {row['overlapping_concepts']}")
                
                # Handle distinct_concepts (converted from gaps_identified)
                if pd.notna(row['distinct_concepts']) and str(row['distinct_concepts']).strip():
                    # Check if it looks like gap analysis or distinct concepts
                    distinct_text = str(row['distinct_concepts']).strip()
                    if any(word in distinct_text.lower() for word in ['gap', 'missing', 'lacking', 'not', 'unable']):
                        explanation_parts.append(f"\n\nGaps Identified: {distinct_text}")
                    else:
                        explanation_parts.append(f"\n\nDistinct Aspects: {distinct_text}")

                # Handle mapping notes (converted from audit_notes)
                if pd.notna(row['mapping_notes']) and str(row['mapping_notes']).strip():
                    explanation_parts.append(f"\n\nAnalyst Notes: {row['mapping_notes']}")

                explanation = ''.join(explanation_parts) if explanation_parts else 'Automated analysis completed - see mapping status for summary.'

                processed_row = {
                    'ID': source_id,
                    'Framework B Control ID': str(row['target_id']),
                    'Framework B Control': str(row['target_text']),
                    'Mapping Status': status_map.get(row['equivalence_type'], 'No Mapping'),
                    'Detailed Mapping Analysis': explanation
                }
                processed_data.append(processed_row)
            
        else:
            # If only 'distinct' mappings exist, create a single 'No Mapping' row
            processed_row = {
                'ID': source_id,
                'Framework B Control ID': 'N/A',
                'Framework B Control': 'No equivalent or overlapping controls found in Framework B.',
                'Mapping Status': 'No Mapping',
                'Detailed Mapping Analysis': 'Automated analysis found no controls in Framework B that share significant conceptual overlap with this Framework A control.'
            }
            processed_data.append(processed_row)
    
    result_df = pd.DataFrame(processed_data)
    logging.info(f"Processed {len(result_df)} individual framework mapping rows.")
    return result_df, None

def create_final_dataframe(merged_df):
    logging.info("Creating final DataFrame for framework mapping report.")
    df = merged_df.copy()
    
    # Fill missing values for Framework A controls that had no matches at all
    df['Mapping Status'] = df['Mapping Status'].fillna('No Mapping')
    df['Detailed Mapping Analysis'] = df['Detailed Mapping Analysis'].fillna('No corresponding controls found in Framework B.')
    df['Framework B Control ID'] = df['Framework B Control ID'].fillna('N/A')
    df['Framework B Control'] = df['Framework B Control'].fillna('No corresponding controls found.')
    
    # Calculate Mapping Score based on Mapping Status
    score_map = {'Direct Mapping': 100, 'Partial Mapping': 50, 'No Mapping': 0}
    df['Mapping Score'] = df['Mapping Status'].map(score_map).fillna(0)
    
    df.insert(0, 'Sr. No.', range(1, len(df) + 1))
    
    # Reorder columns for the final mapping report layout
    final_cols = [
        'Sr. No.', 'Framework A Control ID', 'Framework A Control Domain',
        'Framework A Control Sub-Domain', 'Framework A Control Statement',
        'Framework B Control ID', 'Framework B Control', 'Mapping Score',
        'Detailed Mapping Analysis', 'Mapping Status'
    ]
    
    if not all(col in df.columns for col in final_cols):
        missing = set(final_cols) - set(df.columns)
        msg = f"Internal error: Missing expected columns for final dataframe: {missing}"
        logging.error(msg)
        return None, msg
        
    final = df[final_cols]
    logging.info(f"Final DataFrame created with {len(final)} rows.")
    return final, None

def save_and_format_excel(df, output_path):
    logging.info(f"Saving and formatting Excel mapping report: {output_path}")
    df.to_excel(output_path, index=False, sheet_name="Framework Mapping")
    wb = load_workbook(output_path)
    ws = wb["Framework Mapping"]

    # --- Styling Definitions ---
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid') # Blue header
    
    status_fills = {
        'Direct Mapping': PatternFill('solid', fgColor='C6EFCE'),    # Green
        'Partial Mapping': PatternFill('solid', fgColor='FFEB9C'),  # Yellow
        'No Mapping': PatternFill('solid', fgColor='FFC7CE')        # Red
    }

    # --- Apply Styles ---
    # Header Styling
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')

    # Find column indices once
    header_map = {cell.value: cell.column for cell in ws[1]}
    status_col_idx = header_map.get('Mapping Status')
    framework_a_id_col_idx = header_map.get('Framework A Control ID')

    # Data Rows Formatting
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
        
        # Color-code the 'Mapping Status' column
        if status_col_idx:
            status_cell = ws.cell(row=row[0].row, column=status_col_idx)
            if status_cell.value in status_fills:
                status_cell.fill = status_fills[status_cell.value]

    # --- Vertical Merging for Framework A controls ---
    if framework_a_id_col_idx:
        merge_cols_indices = [
            header_map.get('Framework A Control ID'),
            header_map.get('Framework A Control Domain'),
            header_map.get('Framework A Control Sub-Domain'),
            header_map.get('Framework A Control Statement')
        ]
        
        # Get all data to group by the Framework A control ID
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        
        current_row_num = 2
        for _, group in groupby(data_rows, key=lambda x: x[framework_a_id_col_idx - 1]):
            group_list = list(group)
            num_rows_in_group = len(group_list)
            
            if num_rows_in_group > 1:
                start_row = current_row_num
                end_row = current_row_num + num_rows_in_group - 1
                for col_idx in merge_cols_indices:
                    if col_idx:
                        try:
                            ws.merge_cells(start_row=start_row, start_column=col_idx, end_row=end_row, end_column=col_idx)
                            # Apply center alignment to the merged cell
                            merged_cell = ws.cell(row=start_row, column=col_idx)
                            merged_cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')
                        except Exception as e:
                            logging.warning(f"Could not merge cells in column {col_idx}: {e}")

            current_row_num += num_rows_in_group

    # --- Auto-fit Column Widths ---
    for col in ws.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max(max_length + 2, 15), 60) # Set min 15, max 60 width
        ws.column_dimensions[column_letter].width = adjusted_width

    # Freeze top row
    ws.freeze_panes = 'A2'

    wb.save(output_path)
    logging.info("Excel formatting complete.")

def create_executive_summary(output_path):
    """Creates a comprehensive executive summary sheet for framework mapping analysis."""
    try:
        logging.info("Generating Framework Mapping Executive Summary...")
        wb = load_workbook(output_path)
        ws_data = wb['Framework Mapping']
        
        # Read data into DataFrame
        data = ws_data.values
        cols = next(data)
        df = pd.DataFrame(list(data), columns=cols)

        # Get unique Framework A controls for accurate summary stats
        control_summary = df.drop_duplicates(subset=['Framework A Control ID'])
        
        total_controls = len(control_summary)
        if total_controls == 0:
            logging.warning("No controls found to create an executive summary.")
            return

        direct_mappings = (control_summary['Mapping Status'] == 'Direct Mapping').sum()
        partial_mappings = (control_summary['Mapping Status'] == 'Partial Mapping').sum()
        no_mappings = (control_summary['Mapping Status'] == 'No Mapping').sum()
        
        # Calculate overall mapping coverage
        mapped_controls = direct_mappings + partial_mappings
        mapping_coverage = mapped_controls / total_controls if total_controls > 0 else 0
        
        # Calculate average mapping score
        overall_avg = pd.to_numeric(control_summary['Mapping Score'], errors='coerce').mean()
        
        # Create summary sheet
        if "Executive Summary" in wb.sheetnames:
            del wb["Executive Summary"]
        ws = wb.create_sheet("Executive Summary", 0)

        # --- Summary content and styling ---
        # Title
        ws['A1'] = "Framework Mapping Analysis - Executive Summary"
        ws['A1'].font = Font(size=16, bold=True)
        
        # Key Metrics
        ws['A3'] = "Key Mapping Metrics"
        ws['A3'].font = Font(size=14, bold=True)
        
        ws['A5'] = "Overall Mapping Coverage:"
        ws['B5'] = mapping_coverage
        ws['B5'].number_format = '0.0%'
        
        ws['A6'] = "Average Mapping Score:"
        ws['B6'] = overall_avg / 100
        ws['B6'].number_format = '0.0%'
        
        # Detailed Breakdown
        ws['A8'] = "Mapping Status Breakdown"
        ws['A8'].font = Font(size=14, bold=True)
        
        ws['A10'] = "Mapping Status"
        ws['B10'] = "Count"
        ws['C10'] = "Percentage"
        
        # Header styling
        for cell in [ws['A10'], ws['B10'], ws['C10']]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill('solid', fgColor='E7E6E6')
        
        ws['A11'] = "Direct Mapping"
        ws['B11'] = direct_mappings
        ws['C11'] = direct_mappings / total_controls if total_controls > 0 else 0
        ws['C11'].number_format = '0.0%'
        
        ws['A12'] = "Partial Mapping"
        ws['B12'] = partial_mappings
        ws['C12'] = partial_mappings / total_controls if total_controls > 0 else 0
        ws['C12'].number_format = '0.0%'
        
        ws['A13'] = "No Mapping"
        ws['B13'] = no_mappings
        ws['C13'] = no_mappings / total_controls if total_controls > 0 else 0
        ws['C13'].number_format = '0.0%'
        
        ws['A14'] = "Total Framework A Controls"
        ws['B14'] = total_controls
        ws['C14'] = 1.0
        ws['C14'].number_format = '0.0%'
        
        # Domain breakdown if available
        if 'Framework A Control Domain' in df.columns:
            domain_summary = df.groupby('Framework A Control Domain')['Mapping Status'].value_counts().unstack(fill_value=0)
            if len(domain_summary) > 0:
                ws['A16'] = "Domain-wise Mapping Analysis"
                ws['A16'].font = Font(size=14, bold=True)
                
                # Create domain summary table starting at row 18
                start_row = 18
                domain_summary.reset_index().to_excel(output_path, sheet_name='Executive Summary', 
                                                    startrow=start_row-1, startcol=0, index=False,
                                                    engine='openpyxl')
        
        # Framework Analysis Insights
        ws['A25'] = "Framework Mapping Insights"
        ws['A25'].font = Font(size=14, bold=True)
        
        insights = []
        
        if mapping_coverage >= 0.8:
            insights.append("• High framework alignment - Most Framework A controls have corresponding mappings in Framework B")
        elif mapping_coverage >= 0.5:
            insights.append("• Moderate framework alignment - Significant overlap exists but gaps are present")
        else:
            insights.append("• Limited framework alignment - Substantial differences between frameworks identified")
            
        if direct_mappings > partial_mappings:
            insights.append("• Strong conceptual alignment - Many direct mappings indicate similar control structures")
        elif partial_mappings > no_mappings:
            insights.append("• Partial conceptual alignment - Controls address similar objectives but with different approaches")
        else:
            insights.append("• Distinct framework approaches - Limited conceptual overlap suggests different security philosophies")
            
        # Add insights to worksheet
        current_row = 27
        for insight in insights:
            ws[f'A{current_row}'] = insight
            current_row += 1
        
        # Recommendations
        ws[f'A{current_row + 1}'] = "Recommendations"
        ws[f'A{current_row + 1}'].font = Font(size=14, bold=True)
        
        recommendations = []
        if no_mappings > mapped_controls:
            recommendations.append("• Conduct detailed gap analysis for controls with no mappings")
            recommendations.append("• Consider developing bridge controls to address framework differences")
        if partial_mappings > 0:
            recommendations.append("• Review partial mappings to identify opportunities for enhanced alignment")
        
        recommendations.append("• Use this analysis to inform framework harmonization efforts")
        recommendations.append("• Consider Framework B controls not mapped to identify potential enhancements")
        
        current_row += 3
        for rec in recommendations:
            ws[f'A{current_row}'] = rec
            current_row += 1
        
        # Apply formatting to summary sheet
        for col in ws.columns:
            max_length = 0
            column_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max(max_length + 2, 20), 80)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(output_path)
        logging.info("Framework Mapping Executive Summary created successfully.")
        
    except Exception as e:
        logging.error(f"Could not create executive summary: {e}")

# === Main Orchestrator ===
def generate_report(mapped_data_path: str, framework_path: str, output_path: str) -> tuple[bool, str]:
    """
    Main function to generate the framework mapping report.
    Returns (success: bool, message: str)
    """
    try:
        # 1. Load Framework A definition
        fw_df = load_excel_file(framework_path)
        mapped_fw, err = map_framework_columns(fw_df)
        if err: return False, f"Error processing framework file: {err}"

        # 2. Load overlap analysis data from the analyzer
        mapped_data_df = load_excel_file(mapped_data_path)
        
        # 3. Process the overlap data into report-ready rows
        processed_data, err = process_mapped_data(mapped_data_df)
        if err: return False, f"Error processing overlap analysis data: {err}"

        # 4. Merge framework details with processed results
        # Use a left merge to ensure all Framework A controls are present
        merged_df = pd.merge(mapped_fw, processed_data, on='ID', how='left')

        # 5. Create the final, ordered DataFrame
        final_df, err = create_final_dataframe(merged_df)
        if err: return False, f"Error creating final dataframe: {err}"

        # 6. Save and apply all Excel formatting
        save_and_format_excel(final_df, output_path)
        
        # 7. Create the executive summary sheet
        create_executive_summary(output_path)

        # 8. Generate final summary message
        total = len(final_df.drop_duplicates(subset=['Framework A Control ID']))
        direct_mappings = (final_df['Mapping Status'] == 'Direct Mapping').sum()
        partial_mappings = (final_df['Mapping Status'] == 'Partial Mapping').sum()
        no_mappings = (final_df['Mapping Status'] == 'No Mapping').sum()
        
        mapping_coverage = (direct_mappings + partial_mappings) / total if total > 0 else 0
        
        summary_msg = (
            f"✅ Framework mapping report generated: {output_path}\n"
            f"📊 Framework Mapping Summary ({total} Framework A controls):\n"
            f"  - Direct Mappings:   {direct_mappings}\n"
            f"  - Partial Mappings:  {partial_mappings}\n"
            f"  - No Mappings:       {no_mappings}\n"
            f"  - Coverage Rate:     {mapping_coverage:.1%}"
        )
        return True, summary_msg

    except Exception as e:
        logging.critical(f"A critical error occurred during report generation: {e}", exc_info=True)
        return False, f"Critical Error: {e}"

def main():
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
    
    if len(sys.argv) < 4:
        print("Usage: python framework_mapping_report.py <overlap_analysis.xlsx> <framework_a.xlsx> <output.xlsx>")
        print("\nArguments:")
        print("  overlap_analysis.xlsx: Output from framework overlap analyzer OR old validator script.")
        print("                         Supports both old format ('match_type', 'validation_justification', etc.)")
        print("                         and new format ('equivalence_type', 'mapping_justification', etc.)")
        print("  framework_a.xlsx:      Framework A definition file with columns: 'ID', 'Domain', 'Sub-Domain', 'Controls'")
        print("  output.xlsx:           Path for the final, formatted framework mapping report.")
        print("\nNote: This generates a framework-to-framework mapping analysis report, showing how Framework A controls")
        print("      map to Framework B controls based on conceptual overlap analysis.")
        print("      Backward compatible with old validator output files.")
        sys.exit(1)

    mapped_data_path = sys.argv[1]
    framework_path = sys.argv[2]
    output_path = sys.argv[3]

    success, message = generate_report(mapped_data_path, framework_path, output_path)
    
    if success:
        print(message)
    else:
        print(f"❌ Framework mapping report generation failed: {message}")
        sys.exit(1)

if __name__ == "__main__":
    main()