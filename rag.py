#!/usr/bin/env python3
"""
Maximum Recall Cybersecurity Control Matcher
Target: ~100% Recall for pre-LLM stage
If search misses a match, LLM can't recover it!
"""

import os
import re
import json
import yaml
import hashlib
import numpy as np
import pandas as pd
from typing import Dict, List, Tuple, Set, Optional, Any
from dataclasses import dataclass, field
from collections import defaultdict, Counter
import pickle
from pathlib import Path
import warnings
import sys
import subprocess
import asyncio
import aiohttp
from concurrent.futures import ThreadPoolExecutor
import unicodedata

warnings.filterwarnings('ignore')

# Default configuration with maximum recall settings
DEFAULT_CONFIG = {
    'concept_patterns': {
        'mfa': {
            'patterns': [
                r'\bmfa\b', r'multi[\s-]?factor', r'two[\s-]?factor', r'2fa\b',
                r'authentication.*factor', r'factor.*authentication', r'second[\s-]?factor',
                r'dual[\s-]?factor', r'multi[\s-]?step.*authentication', r'additional.*authentication'
            ],
            'synonyms': [
                'authentication factor', 'second factor', 'dual factor', 'otp', 'one time password',
                'token', 'authenticator', 'pin', 'biometric', 'smart card', 'hardware token', 'software token'
            ],
            'weight': 1.2
        },
        'encryption': {
            'patterns': [
                r'encrypt', r'\btls\b', r'\bssl\b', r'cryptograph', r'data.*transit', r'transit.*encrypt',
                r'data.*rest', r'rest.*encrypt', r'openssh', r'\baes\b', r'cipher', r'key[\s-]?management',
                r'key[\s-]?rotation', r'\bhsm\b', r'hardware.*security.*module', r'database.*encrypt',
                r'field[\s-]?level.*encrypt', r'application[\s-]?layer.*encrypt', r'storage[\s-]?layer.*encrypt'
            ],
            'synonyms': [
                'cipher', 'encode', 'secure transmission', 'data protection', 'cryptographic',
                'encrypted channel', 'secure channel', 'encryption key', 'certificate', 'pki'
            ],
            'weight': 1.2
        },
        'backup': {
            'patterns': [
                r'backup', r'recover', r'restor', r'disaster.*recovery', r'data.*recovery', r'recovery.*data',
                r'replicat', r'snapshot', r'archive', r'failover', r'redundan', r'business.*continuity',
                r'recovery.*point', r'recovery.*time', r'\brpo\b', r'\brto\b'
            ],
            'synonyms': [
                'archive', 'snapshot', 'replicate', 'failover', 'mirror', 'copy', 'redundancy',
                'continuity', 'resilience'
            ],
            'weight': 1.1
        },
        'firewall': {
            'patterns': [
                r'firewall', r'filter.*traffic', r'traffic.*filter', r'network.*filter', r'port.*filter',
                r'packet.*filter', r'stateful.*inspection', r'deep.*packet', r'application.*firewall',
                r'\bwaf\b', r'web.*application.*firewall', r'network.*rule', r'traffic.*rule',
                r'ingress.*rule', r'egress.*rule'
            ],
            'synonyms': [
                'packet filter', 'network security', 'port blocking', 'traffic control',
                'network protection', 'perimeter defense'
            ],
            'weight': 1.0
        },
        'access_control': {
            'patterns': [
                r'access.*control', r'access.*grant', r'grant.*access', r'access.*revok', r'revok.*access',
                r'permission', r'privilege', r'entitlement', r'authorization', r'access.*provision',
                r'provision.*access', r'role[\s-]?based', r'\brbac\b', r'least.*privilege',
                r'need[\s-]?to[\s-]?know', r'segregation.*duties', r'\bsod\b'
            ],
            'synonyms': [
                'authorization', 'entitlement', 'rights management', 'access rights', 'user rights',
                'permissions', 'access management'
            ],
            'weight': 1.1
        },
        'monitoring': {
            'patterns': [
                r'monitor', r'\blog\b', r'logging', r'\bsiem\b', r'detect', r'alert', r'track',
                r'audit.*log', r'event.*log', r'security.*monitor', r'real[\s-]?time.*monitor',
                r'continuous.*monitor', r'log.*aggregat', r'log.*correlat', r'anomaly.*detect',
                r'threat.*detect'
            ],
            'synonyms': [
                'observe', 'watch', 'supervise', 'surveillance', 'tracking', 'auditing', 'logging'
            ],
            'weight': 1.0
        },
        'vulnerability': {
            'patterns': [
                r'vulnerabil', r'patch', r'remediat', r'fix.*security', r'security.*update',
                r'security.*fix', r'vulnerability.*scan', r'security.*scan', r'patch.*management',
                r'vulnerability.*assessment', r'security.*assessment', r'cve', r'zero[\s-]?day', r'exploit'
            ],
            'synonyms': [
                'weakness', 'flaw', 'security hole', 'security gap', 'exposure', 'risk'
            ],
            'weight': 1.0
        },
        'training': {
            'patterns': [
                r'train', r'educat', r'awareness', r'course', r'program.*security', r'security.*program',
                r'security.*awareness', r'awareness.*training', r'phishing.*training', r'security.*education',
                r'user.*training', r'employee.*training'
            ],
            'synonyms': [
                'education', 'learning', 'instruction', 'teaching', 'workshop', 'seminar'
            ],
            'weight': 0.9
        },
        'testing': {
            'patterns': [
                r'test', r'scan', r'assess', r'penetrat', r'pen[\s-]?test', r'vulnerability.*scan',
                r'security.*test', r'security.*assessment', r'audit', r'evaluat', r'security.*audit',
                r'compliance.*audit', r'external.*test', r'internal.*test'
            ],
            'synonyms': [
                'evaluation', 'assessment', 'examination', 'validation', 'verification', 'review'
            ],
            'weight': 0.9
        }
    },
    'critical_concepts': ['mfa', 'encryption', 'backup', 'access_control', 'firewall'],
    'mismatch_penalties': [
        {'source': r'require.*mfa|mfa.*require', 'target': r'test|scan|assess|penetrat', 'penalty': 0.15},
        {'source': r'encrypt.*transit|transit.*encrypt', 'target': r'database|backup|storage|rest', 'penalty': 0.2},
        {'source': r'grant.*access|establish.*process.*grant', 'target': r'expire|revok|terminat', 'penalty': 0.25},
        {'source': r'train.*authentication|authentication.*train', 'target': r'organization.*role|general.*train', 'penalty': 0.2},
        {'source': r'backup|recover', 'target': r'access.*control|authenticat', 'penalty': 0.25},
        {'source': r'require.*vpn|vpn.*authenticat', 'target': r'device.*validat|anti[\s-]?malware', 'penalty': 0.3},
        {'source': r'implement|deploy|configure', 'target': r'test|assess|scan|audit', 'penalty': 0.25}
    ],
    'alignment_boosts': [
        {'source': r'mfa|multi[\s-]?factor', 'target': r'token|factor|pin|authenticat|two[\s-]?factor', 'boost': 0.3},
        {'source': r'encrypt.*transit', 'target': r'tls|ssl|communication|transmission|transfer', 'boost': 0.3},
        {'source': r'backup', 'target': r'backup|recovery|restor|disaster|continuity', 'boost': 0.4},
        {'source': r'firewall', 'target': r'firewall|filter|rule|traffic|packet', 'boost': 0.3},
        {'source': r'train', 'target': r'train|educat|awareness|course|program', 'boost': 0.3},
        {'source': r'vulnerabil', 'target': r'vulnerabil|scan|patch|remediat|update', 'boost': 0.3},
        {'source': r'access.*grant', 'target': r'access.*provision|new.*hire|onboard|create.*account', 'boost': 0.3},
        {'source': r'monitor|log', 'target': r'siem|log|monitor|detect|alert|audit', 'boost': 0.3}
    ],
    'model_weights': {
        'cross_encoder': 0.30,
        'openai': 0.20,
        'sentence_transformer': 0.10,
        'roberta': 0.10,
        'simcse': 0.05,
        'entailment': 0.05,
        'domain': 0.05,
        'tfidf': 0.05,
        'bm25': 0.05,
        'fuzzy': 0.03,
        'concept': 0.02
    },
    'penalty_alpha': 1.5,
    'concept_max_boost': 0.5,
    'candidate_base': 50,
    'candidate_per_concept': 15,
    'score_gap_target': 0.10,
    'async_batch_size': 20,
    'unicode_normalize': True,
    'faiss_nprobe': 50,
    'faiss_neighbors': 100,
    'min_candidates_per_source': 100,
    'use_all_embeddings': True,
    'enable_concept_expansion': True,
    'enable_fuzzy_concept_match': True,
    'max_penalty': 0.5,
    'min_boost': 0.1
}

def check_and_install_requirements():
    """Check if required packages are installed and install if missing."""
    required_packages = {
        'numpy': 'numpy>=1.24.0',
        'pandas': 'pandas>=2.0.0',
        'sklearn': 'scikit-learn>=1.3.0',
        'sentence_transformers': 'sentence-transformers>=2.2.2',
        'rank_bm25': 'rank-bm25>=0.2.2',
        'openai': 'openai>=1.0.0',
        'openpyxl': 'openpyxl>=3.1.0',
        'tqdm': 'tqdm>=4.65.0',
        'torch': 'torch>=2.0.0',
        'transformers': 'transformers>=4.30.0',
        'nltk': 'nltk>=3.8.0',
        'fuzzywuzzy': 'fuzzywuzzy>=0.18.0',
        'Levenshtein': 'python-Levenshtein>=0.20.0',
        'faiss-cpu': 'faiss-cpu>=1.7.0',
        'pyyaml': 'pyyaml>=6.0',
        'jellyfish': 'jellyfish>=0.9.0'
    }
    
    missing_packages = []
    for package, requirement in required_packages.items():
        try:
            __import__(package.replace('-', '_'))
        except ImportError:
            missing_packages.append(requirement)
    
    if missing_packages:
        print("Installing missing packages...")
        for package in missing_packages:
            subprocess.check_call([sys.executable, "-m", "pip", "install", package])
        print("✓ All packages installed successfully")

check_and_install_requirements()

# Imports
from sentence_transformers import SentenceTransformer, CrossEncoder
from rank_bm25 import BM25Okapi
from sklearn.preprocessing import normalize
from sklearn.metrics.pairwise import cosine_similarity
from sklearn.feature_extraction.text import TfidfVectorizer
import openai
from tqdm import tqdm
import faiss
from fuzzywuzzy import fuzz
import jellyfish
import nltk
from nltk.tokenize import word_tokenize
from nltk.corpus import stopwords
from nltk.stem import PorterStemmer
from transformers import pipeline
from openpyxl.styles import Border, Side, Alignment

# Download NLTK data
try:
    nltk.data.find('tokenizers/punkt')
except LookupError:
    nltk.download('punkt', quiet=True)
try:
    nltk.data.find('corpora/stopwords')
except LookupError:
    nltk.download('stopwords', quiet=True)

@dataclass
class Control:
    id: str
    raw_text: str
    processed_text: str
    tokens: List[str]
    stems: List[str]
    expanded_tokens: List[str]
    domain_tags: Set[str]
    domain_scores: Dict[str, float] = field(default_factory=dict)
    keyword_features: Dict[str, int] = field(default_factory=dict)
    concepts: Dict[str, float] = field(default_factory=dict)
    embeddings: Dict[str, np.ndarray] = field(default_factory=dict)

class MaximumRecallMatcher:
    def __init__(self, openai_api_key: str, config_path: Optional[str] = None, cache_dir: str = "./cache"):
        """Initialize the maximum recall matcher."""
        self.openai_api_key = openai_api_key
        openai.api_key = openai_api_key
        self.cache_dir = Path(cache_dir)
        self.cache_dir.mkdir(exist_ok=True)
        
        # Load configuration
        self.config = self._load_config(config_path)
        
        # Initialize NLP tools
        self.stemmer = PorterStemmer()
        self.stop_words = set(stopwords.words('english'))
        
        # Load models
        print("Loading transformer models...")
        self.cross_encoder = CrossEncoder('cross-encoder/ms-marco-MiniLM-L6-v2')
        self.sentence_transformer = SentenceTransformer('all-MiniLM-L6-v2')
        
        # Load additional models if using all embeddings
        if self.config.get('use_all_embeddings', True):
            self.roberta_model = SentenceTransformer('all-roberta-large-v1')
            self.simcse_model = SentenceTransformer('princeton-nlp/sup-simcse-roberta-large')
        
        # Load entailment model if weight > 0
        if self.config['model_weights'].get('entailment', 0) > 0:
            print("Loading entailment model...")
            self.entailment_pipeline = pipeline("zero-shot-classification", 
                                              model="facebook/bart-large-mnli",
                                              device=-1)
        
        # Build synonym dictionary
        self.synonym_dict = self._build_synonym_dict()
        
        # Enhanced boilerplate patterns
        self.boilerplate_patterns = [
            r'^(the\s+)?(organization|company|entity|business|firm)\s+(must|shall|should|will|needs?\s+to)\s+',
            r'^(ensure|verify|confirm|check|validate|assess)\s+that\s+',
            r'^(it\s+is\s+)?(required|mandatory|necessary|essential)\s+(that|to)\s+',
            r'^\d+(\.\d+)*\s*[-:]?\s*',
            r'^[A-Z]{2,}-\d+(\.\d+)*\s*[-:]?\s*',
            r'^(requirement|control|safeguard|measure)\s*[-:]?\s*',
            r'\s*\([a-z]\)\s+',
            r'^\s*[-•]\s+',
            r'^(req|ctrl|sf|ac|cc|cm|ia|ma|mp|pe|pl|pm|ps|sc|si)[-:]?\s*\d+',
            r'^\d+\.\d+\.\d+\s*',
        ]
        
        # TF-IDF vectorizers
        self.tfidf_vectorizer = TfidfVectorizer(
            max_features=5000,
            ngram_range=(1, 3),
            analyzer='word',
            min_df=1
        )
        self.char_tfidf_vectorizer = TfidfVectorizer(
            max_features=3000,
            ngram_range=(3, 6),
            analyzer='char_wb',
            min_df=1
        )
        
        # Load full domain taxonomy
        self.domain_taxonomy = self._load_full_domain_taxonomy()
        self.keyword_to_domains = self._build_keyword_to_domains()
        
        # FAISS indices (will be built during matching)
        self.faiss_indices = {}
    
    def _create_default_config(self, path: str):
        """Create default config file if it doesn't exist."""
        config_path = Path(path)
        config_path.parent.mkdir(parents=True, exist_ok=True)
        
        with open(config_path, 'w') as f:
            yaml.dump(DEFAULT_CONFIG, f, default_flow_style=False, sort_keys=False)
        
        print(f"✓ Created default config at {config_path}")
        return DEFAULT_CONFIG
    
    def _load_config(self, config_path: Optional[str]) -> Dict:
        """Load configuration from YAML file or create default."""
        if config_path:
            path = Path(config_path)
            if not path.exists():
                return self._create_default_config(config_path)
            
            with open(path, 'r') as f:
                user_config = yaml.safe_load(f)
                config = DEFAULT_CONFIG.copy()
                self._deep_merge(config, user_config)
                return config
        else:
            # Create default config in current directory
            default_path = "./config.yaml"
            if not Path(default_path).exists():
                return self._create_default_config(default_path)
            
            with open(default_path, 'r') as f:
                return yaml.safe_load(f)
    
    def _deep_merge(self, base: Dict, update: Dict) -> None:
        """Deep merge update dict into base dict."""
        for key, value in update.items():
            if key in base and isinstance(base[key], dict) and isinstance(value, dict):
                self._deep_merge(base[key], value)
            else:
                base[key] = value
    
    def _build_synonym_dict(self) -> Dict[str, List[str]]:
        """Build comprehensive synonym dictionary."""
        synonym_dict = {}
        
        # Add configured synonyms
        for concept, config in self.config['concept_patterns'].items():
            for synonym in config.get('synonyms', []):
                synonym_lower = synonym.lower()
                if synonym_lower not in synonym_dict:
                    synonym_dict[synonym_lower] = []
                synonym_dict[synonym_lower].append(concept)
        
        # Add extensive technical synonyms
        technical_synonyms = {
            'access': ['privilege', 'permission', 'entitlement', 'authorization', 'right', 'allowance'],
            'virus': ['malware', 'malicious software', 'threat', 'trojan', 'worm', 'ransomware'],
            'authenticate': ['verify identity', 'validate user', 'confirm identity', 'prove identity', 'login'],
            'encrypt': ['cipher', 'encode', 'protect data', 'scramble', 'secure'],
            'monitor': ['track', 'observe', 'watch', 'supervise', 'audit', 'oversee'],
            'test': ['assess', 'evaluate', 'examine', 'validate', 'verify', 'check'],
            'configure': ['setup', 'implement', 'deploy', 'install', 'establish', 'initialize'],
            'user': ['employee', 'personnel', 'staff', 'workforce', 'individual'],
            'network': ['infrastructure', 'system', 'environment', 'architecture'],
            'data': ['information', 'content', 'records', 'files', 'documents'],
            'security': ['protection', 'safeguard', 'defense', 'safety'],
            'control': ['safeguard', 'measure', 'countermeasure', 'protection'],
            'vulnerability': ['weakness', 'flaw', 'exposure', 'risk', 'threat']
        }
        
        for base_word, synonyms in technical_synonyms.items():
            if base_word not in synonym_dict:
                synonym_dict[base_word] = []
            synonym_dict[base_word].extend(synonyms)
        
        return synonym_dict
    
    def _load_full_domain_taxonomy(self) -> Dict[str, List[str]]:
        """Load the complete domain taxonomy."""
        return {
            "Asset Management": [
                "active discovery tool", "android work profile", "app store", "application allowlisting",
                "asset owner", "authorized script", "authorized software", "authorized software library",
                "business purpose", "cloud environment", "decommission date", "department",
                "deployment mechanism", "dhcp logging", "digital signature", "enterprise asset",
                "exception", "hardware address", "install date", "inventory", "iot device",
                "ip address management tool", "license count", "licensed software", "machine name",
                "mdm", "mitigating control", "mobile device", "network address", "network device",
                "network infrastructure", "non-computing device", "passive discovery tool",
                "portable device", "publisher", "residual risk acceptance", "script", "server",
                "software inventory", "software inventory tool", "software library", "supported software",
                "system process", "technical control", "unauthorized asset", "unauthorized software",
                "unsupported software", "url", "use date", "user device", "version", "version control",
                ".dll", ".ocx", ".ps1", ".py", ".so"
            ],
            "Authentication & Access Control": [
                "access control list", "access control review", "access granting", "access permission",
                "access revocation", "access right", "account inventory", "administrator account",
                "audit trail", "authentication system", "authentication system inventory",
                "authorization system", "authorization system inventory", "certificate",
                "dedicated administrator account", "device authentication", "directory service",
                "disable account", "dormant account", "failed authentication attempt", "identity service",
                "least privilege", "local failed authentication attempt", "mfa", "mfa password length",
                "microsoft intune device lock", "need to know", "non-mfa password length",
                "non-privileged account", "password", "password complexity", "port-level access control",
                "privilege", "remote network access", "role-based access control", "sso provider",
                "service account", "service account review", "start date", "stop date", "unique password",
                "user account", "user authentication", "username", "802.1x", "two-factor", "multi-factor",
                "second factor", "authentication factor", "token", "otp", "one-time password"
            ],
            "Data Protection": [
                "access permission", "application", "data access control list", "data classification scheme",
                "data disposal log", "data flow", "data handling", "data inventory", "data management process",
                "data owner", "data processing segmentation", "data retention limit", "data retention timeline",
                "data storage segmentation", "database", "disposal requirement", "dlp", "file system",
                "host-based dlp", "secure disposal", "sensitive data", "sensitive data access log",
                "sensitive data modification log", "service provider data flow", "data loss prevention",
                "information protection", "data security", "data privacy"
            ],
            "Cryptography & Encryption": [
                "application-layer encryption", "apple filevault", "client-side encryption",
                "data at rest encryption", "data in transit encryption", "encrypt data", "linux dm-crypt",
                "openssh", "removable media encryption", "server-side encryption", "storage-layer encryption",
                "tls", "windows bitlocker", "ssl", "https", "certificate", "pki", "public key infrastructure",
                "key management", "key rotation", "hsm", "hardware security module", "aes", "rsa",
                "cryptographic", "cipher", "encryption algorithm", "digital certificate"
            ],
            "Configuration Management & Hardening": [
                "android work profile", "apple configuration profile", "apple configuration profile maxfailedattempts",
                "default account", "default-deny rule", "device lockout", "enterprise-controlled dns server",
                "external dns server", "failed authentication attempt", "firewall", "hardening",
                "host-based firewall", "http", "https", "iac", "infrastructure-as-code",
                "microsoft intune device lock", "operating system firewall", "port-filtering tool",
                "remote wipe", "root account", "secure configuration", "secure configuration process",
                "secure network protocol", "separate enterprise workspace", "session locking", "ssh",
                "telnet", "third-party firewall agent", "trusted dns server", "unnecessary service",
                "virtual firewall", "system hardening", "baseline configuration", "security baseline"
            ],
            "Logging & Monitoring": [
                "abnormal event", "anomaly", "audit log", "audit log management process", "audit log review",
                "authentication event log", "authorization event log", "bash log", "command-line audit log",
                "data creation log", "data disposal log", "destination address", "detailed audit logging",
                "dns query audit log", "event source", "forensic investigation", "log analytics platform",
                "log centralization", "log collection", "log correlation", "log retention", "log review",
                "logging requirement", "powershell log", "potential threat", "remote administrative terminal log",
                "security event alerting", "security event alerting threshold", "service provider log",
                "siem", "source address", "storage capacity", "time source", "time synchronization",
                "timestamp", "url request audit log", "user management event log", "username",
                "vendor-defined event correlation alert", "security information event management",
                "real-time monitoring", "continuous monitoring", "security monitoring"
            ],
            "Network Security": [
                "application layer firewall", "application layer filtering", "architecture diagram",
                "block list", "boundary protection", "browser extension", "browser plugin",
                "category-based filtering", "certificate", "csp service", "dedicated administrative resource",
                "dkim", "dmarc", "dns filtering", "dns filtering service", "edr", "email client extension",
                "email client plugin", "email gateway", "endpoint detection and response", "file type blocking",
                "filtering proxy", "firewall rule", "gateway", "hids", "hips", "home network configuration",
                "host-based intrusion detection solution", "host-based intrusion prevention solution",
                "host-based ips agent", "http", "https", "iac", "infrastructure-as-code", "insecure network",
                "intrusion detection", "intrusion prevention", "least privilege", "logically separated resource",
                "malicious domain", "network aaa", "network availability", "network infrastructure update",
                "network intrusion detection solution", "network intrusion prevention solution",
                "network segmentation", "network system documentation", "network traffic flow log",
                "network-based url filter", "nids", "nips", "physically separated resource",
                "port-level access control", "remote network access", "reputation-based filtering",
                "secure network architecture", "secure network protocol", "spf", "ssh", "supported browser",
                "supported email client", "traffic filtering", "url filter", "vpn", "vpn authentication",
                "wpa2 enterprise", "802.1x", "network access control", "nac", "perimeter security"
            ],
            "Vulnerability Management": [
                "application update", "authenticated scan", "automated patch management", "clear box test",
                "external penetration test", "external vulnerability scan", "internal penetration test",
                "internal vulnerability scan", "opaque box test", "operating system update", "patch management",
                "pen test", "penetration test", "penetration test finding remediation",
                "penetration testing program", "reconnaissance", "remediation process",
                "risk-based remediation strategy", "unauthenticated scan", "vulnerability management process",
                "vulnerability remediation", "vulnerability scan", "security patch", "hotfix", "security update",
                "cve", "common vulnerabilities and exposures", "zero day", "exploit", "security assessment"
            ],
            "Incident Response": [
                "communication plan", "contact information", "cyber insurance provider contact",
                "data breach", "government agency contact", "incident communication plan",
                "incident contact information", "incident handling", "incident handling process",
                "incident recovery effort", "incident reporting process", "incident response coordination",
                "incident response exercise", "incident response plan", "isac", "isac partner",
                "law enforcement contact", "lesson learned", "post-incident review", "privacy incident",
                "security incident threshold", "breach notification", "incident management",
                "computer security incident response team", "csirt", "security operations center", "soc"
            ],
            "Software Development Security": [
                "application penetration testing", "application security testing", "attack surface minimization",
                "auditing component", "authenticated penetration testing", "bill of materials", "bom",
                "business logic vulnerability", "developer training", "dynamic analysis", "encryption component",
                "encryption algorithm", "error checking", "framework", "hardening configuration template",
                "identity management component", "least privilege", "library", "logging component",
                "non-production environment", "owasp top 10", "production environment", "root cause analysis",
                "secure application design standard", "secure application development",
                "secure application development process", "secure code training", "secure coding practice",
                "secure design principle", "severity rating", "software vulnerability report", "static analysis",
                "third-party code security", "third-party component", "third-party component inventory",
                "threat modeling", "trusted third-party component", "vetted module", "vetted service",
                "vulnerability handling policy", "vulnerability management", "vulnerability tracking system",
                "sast", "dast", "code review", "security testing"
            ],
            "Human Resources Security & Security Awareness Training": [
                "authentication best practice training", "bec training", "business email compromise training",
                "clear desk practice", "clear screen practice", "credential management training",
                "home network configuration training", "incident recognition training",
                "incident reporting training", "insecure network training", "mfa training",
                "owasp top 10 training", "password composition training", "phishing training",
                "pretexting training", "role-specific training", "security awareness program",
                "secure system administration course", "sensitive data handling training",
                "social engineering attack training", "software patch failure training",
                "automated process failure training", "tailgating training", "unintentional data exposure training",
                "security education", "awareness campaign", "security culture"
            ],
            "Third-Party / Service Provider Management": [
                "account deactivation", "dark web monitoring", "data flow termination",
                "data disposal commitment", "data encryption requirement", "incident notification requirement",
                "pci aoc", "pci attestation of compliance", "questionnaire", "security requirement",
                "service provider assessment", "service provider classification", "service provider contract",
                "service provider decommissioning", "service provider inventory",
                "service provider management policy", "service provider monitoring", "soc 2 report",
                "secure data disposal", "vendor management", "supplier security", "third party risk"
            ],
            "Physical Security": [
                "physical premise control", "physically separated resource", "tailgating training",
                "physical access control", "badge access", "security guard", "cctv", "surveillance",
                "secure facility", "data center security", "environmental controls"
            ],
            "Malware Defenses": [
                "anti-exploitation feature", "anti-malware signature file", "anti-malware software",
                "attachment scanning", "autoplay", "autorun", "behavior-based anti-malware",
                "data execution prevention", "dep", "email server anti-malware", "gatekeeper",
                "microsoft data execution prevention", "removable media scan", "sandboxing", "sip",
                "system integrity protection", "wdeg", "windows defender exploit guard",
                "antivirus", "anti-virus", "endpoint protection", "epp", "malware detection"
            ],
            "Data Recovery": [
                "automated backup", "backup data security", "backup recovery test", "cloud backup",
                "data recovery process", "isolated recovery data", "off-site backup", "offline backup",
                "recovery data protection", "disaster recovery", "business continuity", "rpo", "rto",
                "recovery point objective", "recovery time objective", "backup strategy", "restore procedure"
            ]
        }
    
    def _build_keyword_to_domains(self) -> Dict[str, Set[str]]:
        """Build reverse index from keywords to domains."""
        keyword_to_domains = defaultdict(set)
        for domain, keywords in self.domain_taxonomy.items():
            for keyword in keywords:
                keyword_to_domains[keyword.lower()].add(domain)
        return dict(keyword_to_domains)
    
    def _normalize_unicode(self, text: str) -> str:
        """Normalize unicode characters to ASCII."""
        if self.config.get('unicode_normalize', True):
            text = unicodedata.normalize('NFKD', text)
            text = text.encode('ascii', 'ignore').decode('ascii')
        return text
    
    def _expand_with_synonyms(self, tokens: List[str]) -> List[str]:
        """Expand tokens with comprehensive synonyms."""
        expanded = list(tokens)
        
        for token in tokens:
            token_lower = token.lower()
            
            # Direct synonyms
            if token_lower in self.synonym_dict:
                expanded.extend(self.synonym_dict[token_lower])
            
            # Reverse lookup
            for base_word, synonyms in self.synonym_dict.items():
                if token_lower in [s.lower() for s in synonyms]:
                    expanded.append(base_word)
                    expanded.extend(synonyms)
        
        # Add concept-based expansions
        if self.config.get('enable_concept_expansion', True):
            for concept, config in self.config['concept_patterns'].items():
                for pattern in config['patterns']:
                    if re.search(pattern, token_lower):
                        expanded.extend(config.get('synonyms', []))
        
        return list(set(expanded))
    
    def _extract_concepts(self, text: str) -> Dict[str, float]:
        """Extract concept scores with fuzzy matching."""
        text_lower = text.lower()
        concepts = {}
        
        for concept, config in self.config['concept_patterns'].items():
            score = 0.0
            weight = config.get('weight', 1.0)
            
            # Pattern matching
            for pattern in config['patterns']:
                matches = len(re.findall(pattern, text_lower))
                score += matches * weight
            
            # Synonym matching
            for synonym in config.get('synonyms', []):
                if synonym.lower() in text_lower:
                    score += 0.5 * weight
            
            # Fuzzy concept matching
            if self.config.get('enable_fuzzy_concept_match', True):
                # Check for partial matches
                if concept in text_lower:
                    score += 0.3 * weight
            
            if score > 0:
                concepts[concept] = min(score, 3.0)
        
        return concepts
    
    def _calculate_soft_penalty(self, source_text: str, target_text: str) -> float:
        """Calculate soft logistic penalty with cap."""
        source_lower = source_text.lower()
        target_lower = target_text.lower()
        
        total_penalty = 0.0
        alpha = self.config.get('penalty_alpha', 1.5)
        
        for penalty_config in self.config['mismatch_penalties']:
            if re.search(penalty_config['source'], source_lower) and \
               re.search(penalty_config['target'], target_lower):
                total_penalty += penalty_config['penalty']
        
        # Cap maximum penalty
        total_penalty = min(total_penalty, self.config.get('max_penalty', 0.5))
        
        # Soft logistic penalty
        return 1.0 / (1.0 + np.exp(alpha * total_penalty))
    
    def _calculate_alignment_boost(self, source_text: str, target_text: str) -> float:
        """Calculate alignment boost with minimum guarantee."""
        source_lower = source_text.lower()
        target_lower = target_text.lower()
        
        total_boost = 0.0
        
        for boost_config in self.config['alignment_boosts']:
            if re.search(boost_config['source'], source_lower) and \
               re.search(boost_config['target'], target_lower):
                total_boost += boost_config['boost']
        
        # Apply minimum boost if there's any alignment
        if total_boost > 0:
            total_boost = max(total_boost, self.config.get('min_boost', 0.1))
        
        return min(total_boost, self.config.get('concept_max_boost', 0.5))
    
    def _advanced_text_processing(self, text: str) -> Control:
        """Process text with maximum feature extraction."""
        # Unicode normalization
        text = self._normalize_unicode(text)
        
        # Clean text
        cleaned = text.lower().strip()
        
        # Remove boilerplate
        for pattern in self.boilerplate_patterns:
            cleaned = re.sub(pattern, '', cleaned, flags=re.IGNORECASE)
        
        # Tokenize
        tokens = word_tokenize(cleaned)
        
        # Remove stopwords and stem
        filtered_tokens = [token for token in tokens if token not in self.stop_words and token.isalnum()]
        stems = [self.stemmer.stem(token) for token in filtered_tokens]
        
        # Aggressive synonym expansion
        expanded_tokens = self._expand_with_synonyms(filtered_tokens)
        
        # Extract concepts with fuzzy matching
        concepts = self._extract_concepts(text)
        
        # Extract domains
        domain_tags = set()
        domain_scores = defaultdict(float)
        keyword_features = defaultdict(int)
        
        cleaned_lower = cleaned.lower()
        for keyword, domains in self.keyword_to_domains.items():
            if keyword in cleaned_lower:
                for domain in domains:
                    domain_tags.add(domain)
                    domain_scores[domain] += 1.0
                    keyword_features[keyword] += cleaned_lower.count(keyword)
        
        # Normalize domain scores
        if domain_scores:
            max_score = max(domain_scores.values())
            domain_scores = {k: v/max_score for k, v in domain_scores.items()}
        
        return Control(
            id='',
            raw_text=text,
            processed_text=cleaned,
            tokens=tokens,
            stems=stems,
            expanded_tokens=expanded_tokens,
            domain_tags=domain_tags,
            domain_scores=dict(domain_scores),
            keyword_features=dict(keyword_features),
            concepts=concepts,
            embeddings={}
        )
    
    async def _get_openai_embeddings_async(self, texts: List[str]) -> List[np.ndarray]:
        """Get OpenAI embeddings asynchronously with retry."""
        embeddings = []
        batch_size = self.config.get('async_batch_size', 20)
        
        async with aiohttp.ClientSession() as session:
            for i in range(0, len(texts), batch_size):
                batch = texts[i:i + batch_size]
                
                # Retry logic
                for attempt in range(3):
                    try:
                        headers = {
                            'Authorization': f'Bearer {self.openai_api_key}',
                            'Content-Type': 'application/json'
                        }
                        
                        data = {
                            'model': 'text-embedding-3-large',
                            'input': batch
                        }
                        
                        async with session.post(
                            'https://api.openai.com/v1/embeddings',
                            headers=headers,
                            json=data
                        ) as response:
                            result = await response.json()
                            
                            for embedding_data in result['data']:
                                embeddings.append(np.array(embedding_data['embedding']))
                            break
                    except Exception as e:
                        if attempt == 2:
                            print(f"Error getting embeddings: {e}")
                            # Return zero embeddings as fallback
                            embeddings.extend([np.zeros(3072) for _ in batch])
                        else:
                            await asyncio.sleep(2 ** attempt)
        
        return embeddings
    
    def _get_embeddings_batch(self, texts: List[str], models: List[str]) -> Dict[str, np.ndarray]:
        """Get embeddings from multiple models with fallback."""
        n_texts = len(texts)
        embeddings = {}
        
        # OpenAI embeddings
        if 'openai' in models:
            print("Getting OpenAI embeddings...")
            try:
                if len(texts) > 50:
                    loop = asyncio.new_event_loop()
                    asyncio.set_event_loop(loop)
                    openai_embs = loop.run_until_complete(self._get_openai_embeddings_async(texts))
                    embeddings['openai'] = np.array(openai_embs)
                else:
                    response = openai.embeddings.create(
                        model="text-embedding-3-large",
                        input=texts
                    )
                    embeddings['openai'] = np.array([d.embedding for d in response.data])
            except Exception as e:
                print(f"OpenAI embedding error: {e}")
                embeddings['openai'] = np.random.randn(n_texts, 3072)
        
        # Sentence transformer
        if 'sentence_transformer' in models:
            print("Getting sentence transformer embeddings...")
            embeddings['sentence_transformer'] = self.sentence_transformer.encode(texts, show_progress_bar=True)
        
        # Additional models if configured
        if self.config.get('use_all_embeddings', True):
            if 'roberta' in models and hasattr(self, 'roberta_model'):
                print("Getting RoBERTa embeddings...")
                embeddings['roberta'] = self.roberta_model.encode(texts, show_progress_bar=True)
            
            if 'simcse' in models and hasattr(self, 'simcse_model'):
                print("Getting SimCSE embeddings...")
                embeddings['simcse'] = self.simcse_model.encode(texts, show_progress_bar=True)
        
        return embeddings
    
    def _build_faiss_index(self, embeddings: np.ndarray) -> faiss.Index:
        """Build FAISS index with optimized parameters."""
        dimension = embeddings.shape[1]
        
        # Normalize embeddings
        normalized = embeddings / (np.linalg.norm(embeddings, axis=1, keepdims=True) + 1e-10)
        
        # Use IndexFlatIP for maximum recall
        index = faiss.IndexFlatIP(dimension)
        index.add(normalized.astype('float32'))
        
        return index
    
    def _calculate_entailment_score(self, text1: str, text2: str) -> float:
        """Calculate entailment score if enabled."""
        if not hasattr(self, 'entailment_pipeline'):
            return 0.0
        
        try:
            result = self.entailment_pipeline(
                text1[:500],
                candidate_labels=[text2[:500]],
                hypothesis_template="This text means: {}"
            )
            return result['scores'][0]
        except:
            return 0.0
    
    def _calculate_advanced_fuzzy_scores(self, text1: str, text2: str) -> Dict[str, float]:
        """Calculate comprehensive fuzzy string similarity scores."""
        # Limit text length for expensive operations
        t1 = text1[:1000]
        t2 = text2[:1000]
        
        scores = {
            'fuzz_ratio': fuzz.ratio(t1, t2) / 100.0,
            'fuzz_partial': fuzz.partial_ratio(t1, t2) / 100.0,
            'fuzz_token_sort': fuzz.token_sort_ratio(t1, t2) / 100.0,
            'fuzz_token_set': fuzz.token_set_ratio(t1, t2) / 100.0,
        }
        
        # Add Jaro-Winkler and Damerau-Levenshtein for shorter texts
        if len(t1) < 200 and len(t2) < 200:
            scores['jaro_winkler'] = jellyfish.jaro_winkler_similarity(t1, t2)
            scores['damerau_levenshtein'] = 1.0 - (jellyfish.damerau_levenshtein_distance(t1[:100], t2[:100]) / 
                                                   max(len(t1[:100]), len(t2[:100])))
        else:
            scores['jaro_winkler'] = scores['fuzz_ratio']
            scores['damerau_levenshtein'] = scores['fuzz_ratio']
        
        return scores
    
    def _ensure_minimum_candidates(self, current_candidates: Set[int], n_target: int, 
                                  min_required: int) -> Set[int]:
        """Ensure we have minimum number of candidates."""
        if len(current_candidates) >= min_required:
            return current_candidates
        
        # Add random candidates to meet minimum
        all_indices = set(range(n_target))
        remaining = all_indices - current_candidates
        
        n_to_add = min_required - len(current_candidates)
        if remaining:
            additional = np.random.choice(list(remaining), 
                                        min(n_to_add, len(remaining)), 
                                        replace=False)
            current_candidates.update(additional)
        
        return current_candidates
    
    def _critical_concept_fallback(self, source_ctrl: Control, target_controls: List[Control], 
                                  current_matches: List[int]) -> List[int]:
        """Ensure critical concepts are in top matches."""
        critical_concepts = set(self.config['critical_concepts'])
        source_critical = set(source_ctrl.concepts.keys()) & critical_concepts
        
        if not source_critical:
            return current_matches
        
        # Check top 5 for critical concepts
        has_critical = False
        for idx in current_matches[:5]:
            target_critical = set(target_controls[idx].concepts.keys()) & source_critical
            if target_critical:
                has_critical = True
                break
        
        if not has_critical:
            # Find ALL targets with critical concepts
            critical_candidates = []
            for j, target_ctrl in enumerate(target_controls):
                if j in current_matches[:10]:
                    continue
                
                target_critical = set(target_ctrl.concepts.keys()) & source_critical
                if target_critical:
                    overlap_score = len(target_critical) / len(source_critical)
                    critical_candidates.append((j, overlap_score))
            
            # Sort by overlap and insert top ones
            critical_candidates.sort(key=lambda x: x[1], reverse=True)
            
            # Insert multiple critical matches, avoiding duplicates
            insert_positions = [2, 3, 4]  # Positions 3-5
            inserted_indices = set()
            
            for pos, (idx, _) in zip(insert_positions, critical_candidates[:3]):
                if idx not in inserted_indices and idx not in current_matches:
                    if pos < len(current_matches):
                        current_matches.insert(pos, idx)
                        inserted_indices.add(idx)
        
        return current_matches
    
    def load_excel_controls(self, filepath: str, sheet_name: str = None) -> List[Dict]:
        """Load controls from Excel file using 'Controls' and 'ID' columns."""
        df = pd.read_excel(filepath, sheet_name=sheet_name)
        
        if isinstance(df, dict):
            print(f"⚠ {filepath} has multiple sheets. Using the first one.")
            df = list(df.values())[0]
        
        # Check for required columns
        if 'Controls' not in df.columns or 'ID' not in df.columns:
            raise ValueError(f"Excel file must have 'Controls' and 'ID' columns. Found: {list(df.columns)}")
        
        controls = []
        for idx, row in df.iterrows():
            control_text = str(row['Controls']).strip()
            control_id = str(row['ID']).strip()
            
            if control_text and control_text.lower() != 'nan':
                controls.append({
                    'id': control_id,
                    'raw_text': control_text
                })
        
        return controls
    
    def match_controls(self, source_controls: List[Dict], target_controls: List[Dict], top_k: int = 5) -> pd.DataFrame:
        """Main matching pipeline optimized for maximum recall."""
        print("=" * 80)
        print("MAXIMUM RECALL MATCHER - TARGET: ~100% RECALL")
        print(f"Finding top {top_k} matches with {self.config['min_candidates_per_source']} minimum candidates")
        print("=" * 80)
        
        # Process controls
        print("\n[1/8] Processing Controls with Synonym Expansion")
        source_processed = []
        target_processed = []
        
        for ctrl in tqdm(source_controls, desc="Processing source controls"):
            control = self._advanced_text_processing(ctrl['raw_text'])
            control.id = ctrl['id']
            source_processed.append(control)
        
        for ctrl in tqdm(target_controls, desc="Processing target controls"):
            control = self._advanced_text_processing(ctrl['raw_text'])
            control.id = ctrl['id']
            target_processed.append(control)
        
        # Get embeddings
        print("\n[2/8] Generating Multi-Model Embeddings")
        source_texts = [c.processed_text for c in source_processed]
        target_texts = [c.processed_text for c in target_processed]
        all_texts = source_texts + target_texts
        
        # Use all available models
        models = ['openai', 'sentence_transformer']
        if self.config.get('use_all_embeddings', True):
            models.extend(['roberta', 'simcse'])
        
        embeddings = self._get_embeddings_batch(all_texts, models)
        
        # Split embeddings
        n_source = len(source_texts)
        source_embeddings = {m: emb[:n_source] for m, emb in embeddings.items()}
        target_embeddings = {m: emb[n_source:] for m, emb in embeddings.items()}
        
        # Build FAISS indices
        print("\n[3/8] Building FAISS Indices for Fast Retrieval")
        faiss_indices = {}
        for model in models[:2]:  # Use top 2 models for FAISS
            if model in target_embeddings:
                faiss_indices[model] = self._build_faiss_index(target_embeddings[model])
        
        # TF-IDF features
        print("\n[4/8] Building TF-IDF Features")
        word_tfidf = self.tfidf_vectorizer.fit_transform(all_texts)
        char_tfidf = self.char_tfidf_vectorizer.fit_transform(all_texts)
        
        word_tfidf_source = word_tfidf[:n_source]
        word_tfidf_target = word_tfidf[n_source:]
        char_tfidf_source = char_tfidf[:n_source]
        char_tfidf_target = char_tfidf[n_source:]
        
        # BM25 with expanded tokens
        print("\n[5/8] Building BM25 Index with Expanded Tokens")
        target_expanded = [c.expanded_tokens for c in target_processed]
        bm25 = BM25Okapi(target_expanded)
        
        # Pre-calculate penalties and boosts
        print("\n[6/8] Pre-calculating Penalties and Boosts")
        n_target = len(target_processed)
        penalty_matrix = np.ones((n_source, n_target))
        boost_matrix = np.zeros((n_source, n_target))
        
        for i in range(n_source):
            for j in range(n_target):
                penalty_matrix[i, j] = self._calculate_soft_penalty(
                    source_processed[i].raw_text,
                    target_processed[j].raw_text
                )
                boost_matrix[i, j] = self._calculate_alignment_boost(
                    source_processed[i].raw_text,
                    target_processed[j].raw_text
                )
        
        # Multi-stage candidate selection
        print(f"\n[7/8] Multi-Stage Candidate Selection")
        results = []
        
        for i, source_ctrl in enumerate(tqdm(source_processed, desc="Finding matches")):
            # Dynamic candidate count based on concepts
            n_concepts = len(source_ctrl.concepts)
            n_critical = len(set(source_ctrl.concepts.keys()) & set(self.config['critical_concepts']))
            
            base_candidates = self.config['candidate_base']
            per_concept = self.config['candidate_per_concept']
            min_candidates = self.config.get('min_candidates_per_source', 100)
            
            target_candidates = max(
                min_candidates,
                base_candidates + per_concept * n_critical
            )
            
            # Stage 1: FAISS retrieval (get many candidates)
            faiss_candidates = set()
            faiss_k = self.config.get('faiss_neighbors', 100)
            
            for model, index in faiss_indices.items():
                if model in source_embeddings:
                    source_emb = source_embeddings[model][i:i+1]
                    source_emb_norm = source_emb / (np.linalg.norm(source_emb) + 1e-10)
                    
                    D, I = index.search(source_emb_norm.astype('float32'), min(faiss_k, n_target))
                    faiss_candidates.update(I[0])
            
            # Stage 2: BM25 retrieval with expanded tokens
            bm25_scores = bm25.get_scores(source_ctrl.expanded_tokens)
            bm25_top = min(80, n_target)  # Get more BM25 candidates
            bm25_candidates = set(np.argsort(bm25_scores)[-bm25_top:][::-1])
            
            # Stage 3: Concept-based retrieval
            concept_candidates = set()
            if source_ctrl.concepts:
                concept_scores = []
                for j, target_ctrl in enumerate(target_processed):
                    # Calculate concept overlap score
                    common = set(source_ctrl.concepts.keys()) & set(target_ctrl.concepts.keys())
                    if common:
                        score = sum(source_ctrl.concepts[c] * target_ctrl.concepts.get(c, 0) for c in common)
                        concept_scores.append((j, score))
                
                # Get top concept matches
                concept_scores.sort(key=lambda x: x[1], reverse=True)
                concept_candidates.update([idx for idx, _ in concept_scores[:50]])
            
            # Stage 4: Domain-based retrieval
            domain_candidates = set()
            if source_ctrl.domain_tags:
                for j, target_ctrl in enumerate(target_processed):
                    if len(source_ctrl.domain_tags & target_ctrl.domain_tags) >= 1:
                        domain_candidates.add(j)
            
            # Combine all candidates with union
            all_candidates = faiss_candidates | bm25_candidates | concept_candidates | domain_candidates
            
            # Ensure minimum candidates
            all_candidates = self._ensure_minimum_candidates(all_candidates, n_target, target_candidates)
            
            # Stage 5: Deep scoring for all candidates
            candidate_scores = []
            
            for j in all_candidates:
                target_ctrl = target_processed[j]
                
                # Cross-encoder score
                ce_score = self.cross_encoder.predict([(source_ctrl.raw_text, target_ctrl.raw_text)])[0]
                
                # Embedding similarities
                emb_scores = {}
                for model in models:
                    if model in source_embeddings and model in target_embeddings:
                        sim = cosine_similarity(
                            source_embeddings[model][i:i+1],
                            target_embeddings[model][j:j+1]
                        )[0, 0]
                        emb_scores[model] = max(0, sim)  # Ensure non-negative
                
                # Fuzzy scores
                fuzzy_scores = self._calculate_advanced_fuzzy_scores(
                    source_ctrl.processed_text,
                    target_ctrl.processed_text
                )
                
                # TF-IDF similarities
                word_tfidf_sim = cosine_similarity(
                    word_tfidf_source[i:i+1],
                    word_tfidf_target[j:j+1]
                )[0, 0]
                char_tfidf_sim = cosine_similarity(
                    char_tfidf_source[i:i+1],
                    char_tfidf_target[j:j+1]
                )[0, 0]
                
                # Concept similarity
                concept_sim = 0.0
                if source_ctrl.concepts and target_ctrl.concepts:
                    common_concepts = set(source_ctrl.concepts.keys()) & set(target_ctrl.concepts.keys())
                    if common_concepts:
                        # Weighted concept similarity
                        concept_sim = sum(
                            min(source_ctrl.concepts[c], target_ctrl.concepts[c]) /
                            max(source_ctrl.concepts[c], target_ctrl.concepts[c])
                            for c in common_concepts
                        ) / len(common_concepts)
                
                # Domain similarity
                domain_sim = 0.0
                if source_ctrl.domain_tags and target_ctrl.domain_tags:
                    common_domains = source_ctrl.domain_tags & target_ctrl.domain_tags
                    all_domains = source_ctrl.domain_tags | target_ctrl.domain_tags
                    domain_sim = len(common_domains) / len(all_domains) if all_domains else 0.0
                
                # BM25 normalized score
                bm25_norm = bm25_scores[j] / (max(bm25_scores) + 1e-10)
                
                # Calculate weighted score
                weights = self.config['model_weights']
                
                base_score = (
                    weights['cross_encoder'] * ce_score +
                    weights['openai'] * emb_scores.get('openai', 0) +
                    weights['sentence_transformer'] * emb_scores.get('sentence_transformer', 0) +
                    weights.get('roberta', 0) * emb_scores.get('roberta', 0) +
                    weights.get('simcse', 0) * emb_scores.get('simcse', 0) +
                    weights['domain'] * domain_sim +
                    weights['tfidf'] * (word_tfidf_sim + char_tfidf_sim) / 2 +
                    weights['bm25'] * bm25_norm +
                    weights['fuzzy'] * fuzzy_scores['fuzz_token_set'] +
                    weights['concept'] * concept_sim
                )
                
                # Apply penalty and boost
                final_score = base_score * penalty_matrix[i, j] + boost_matrix[i, j]
                
                # Extra boost for critical concept matches
                if source_ctrl.concepts and target_ctrl.concepts:
                    critical_overlap = (set(source_ctrl.concepts.keys()) & 
                                      set(target_ctrl.concepts.keys()) & 
                                      set(self.config['critical_concepts']))
                    if critical_overlap:
                        final_score += 0.1 * len(critical_overlap)
                
                candidate_scores.append({
                    'idx': j,
                    'score': final_score,
                    'ce_score': ce_score,
                    'concept_overlap': len(set(source_ctrl.concepts.keys()) & set(target_ctrl.concepts.keys()))
                })
            
            # Sort by score
            candidate_scores.sort(key=lambda x: x['score'], reverse=True)
            
            # Get indices
            top_indices = [c['idx'] for c in candidate_scores]
            
            # Apply critical concept fallback
            top_indices = self._critical_concept_fallback(source_ctrl, target_processed, top_indices)
            
            # Remove duplicates while preserving order
            seen = set()
            unique_indices = []
            for idx in top_indices:
                if idx not in seen:
                    seen.add(idx)
                    unique_indices.append(idx)
            
            # Take top k with safety margin
            for rank, idx in enumerate(unique_indices[:top_k]):
                target_ctrl = target_processed[idx]
                
                results.append({
                    'source_id': source_ctrl.id,
                    'source_text': source_ctrl.raw_text,
                    'target_id': target_ctrl.id,
                    'target_text': target_ctrl.raw_text,
                    'rank': rank + 1
                })
        
        print(f"\n[8/8] Generated {len(results)} total matches")
        
        return pd.DataFrame(results)
    
    def run_pipeline(self, source_excel: str, target_excel: str, output_dir: str = "./output", 
                    top_k: int = 5, config_path: Optional[str] = None):
        """Run the complete pipeline."""
        output_path = Path(output_dir)
        output_path.mkdir(exist_ok=True)
        
        # Save config
        with open(output_path / 'config.yaml', 'w') as f:
            yaml.dump(self.config, f, default_flow_style=False, sort_keys=False)
        
        print("\n" + "=" * 80)
        print("LOADING CONTROL DATA")
        print("=" * 80)
        
        source_controls = self.load_excel_controls(source_excel)
        target_controls = self.load_excel_controls(target_excel)
        
        print(f"✓ Loaded {len(source_controls)} source controls")
        print(f"✓ Loaded {len(target_controls)} target controls")
        
        # Run matching
        results_df = self.match_controls(source_controls, target_controls, top_k=top_k)
        
        # Save outputs with formatting
        # Save CSV
        results_df.to_csv(output_path / f"max_recall_top{top_k}.csv", index=False)
        print(f"✓ Saved results to {output_path / f'max_recall_top{top_k}.csv'}")
        
        # Save formatted Excel
        with pd.ExcelWriter(output_path / f"max_recall_top{top_k}.xlsx", engine='openpyxl') as writer:
            results_df.to_excel(writer, sheet_name='Matches', index=False)
            
            # Get worksheet
            worksheet = writer.sheets['Matches']
            
            # Define border style
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Apply formatting
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, 
                                         min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    # Apply border
                    cell.border = thin_border
                    # Apply wrap text
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            # Set column widths (25 characters, not pixels)
            for column in ['A', 'B', 'C', 'D']:
                worksheet.column_dimensions[column].width = 25
        
        print(f"✓ Saved formatted Excel to {output_path / f'max_recall_top{top_k}.xlsx'}")
        
        # Summary report
        print("\n" + "=" * 80)
        print("MATCHING SUMMARY")
        print("=" * 80)
        print(f"Total Matches: {len(results_df)}")
        print(f"Unique Source Controls: {results_df['source_id'].nunique()}")
        print(f"Unique Target Controls: {results_df['target_id'].nunique()}")
        print("=" * 80)
        
        return results_df
    
    def _calculate_critical_coverage(self, results_df: pd.DataFrame) -> float:
        """Calculate critical concept coverage."""
        critical_concepts = set(self.config['critical_concepts'])
        covered = 0
        total = 0
        
        for source_id in results_df['source_id'].unique():
            source_top5 = results_df[
                (results_df['source_id'] == source_id) & 
                (results_df['rank'] <= 5)
            ]
            
            # Check if source has critical concepts
            source_concepts = set()
            if not source_top5.empty:
                first_row = source_top5.iloc[0]
                if first_row['source_concepts']:
                    source_concepts = set(first_row['source_concepts'].split(', '))
            
            source_critical = source_concepts & critical_concepts
            if source_critical:
                total += 1
                
                # Check if any top-5 match has the critical concept
                for _, row in source_top5.iterrows():
                    if row['target_concepts']:
                        target_concepts = set(row['target_concepts'].split(', '))
                        if source_critical & target_concepts:
                            covered += 1
                            break
        
        return covered / total if total > 0 else 1.0

def main():
    if len(sys.argv) < 3:
        print("\n" + "=" * 80)
        print("MAXIMUM RECALL CYBERSECURITY CONTROL MATCHER")
        print("=" * 80)
        print("\nUsage: python matcher.py <source_excel> <target_excel> [output_dir] [top_k] [config.yaml]")
        print("\nExample:")
        print("  python matcher.py framework.xlsx soc2.xlsx ./results 5")
        print("  python matcher.py framework.xlsx soc2.xlsx ./results 5 custom_config.yaml")
        print("\nNote: Excel files must have 'Controls' and 'ID' columns")
        print("=" * 80)
        sys.exit(1)
    
    api_key = os.environ.get("OPENAI_API_KEY")
    if not api_key:
        print("\n❌ Error: OpenAI API key not found!")
        print("Set: export OPENAI_API_KEY='your-key-here'")
        sys.exit(1)
    
    source_file = sys.argv[1]
    target_file = sys.argv[2]
    output_dir = sys.argv[3] if len(sys.argv) > 3 else "./recall_output"
    top_k = int(sys.argv[4]) if len(sys.argv) > 4 else 5
    config_path = sys.argv[5] if len(sys.argv) > 5 else None
    
    if not os.path.exists(source_file):
        print(f"❌ Source file not found: {source_file}")
        sys.exit(1)
    
    if not os.path.exists(target_file):
        print(f"❌ Target file not found: {target_file}")
        sys.exit(1)
    
    matcher = MaximumRecallMatcher(api_key, config_path)
    matcher.run_pipeline(source_file, target_file, output_dir, top_k, config_path)

if __name__ == "__main__":
    main()